#!/bin/bash
from datetime import datetime
import os
from flask import Flask, jsonify, request, make_response, redirect, flash, Markup, url_for, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_httpauth import HTTPBasicAuth
from flask_admin import Admin, BaseView, expose
from flask_admin.contrib.sqla import ModelView
from sqlalchemy import exc, func, and_
import codecs
from PIL import Image
from flask_swagger_ui import get_swaggerui_blueprint
from predictor import process_photo, process_video, stop_processing
from flask_babel import Babel
from log import log
from openpyxl import load_workbook

app = Flask(__name__)
app.config.from_pyfile('config.py')
db = SQLAlchemy(app)
auth = HTTPBasicAuth()
SWAGGER_URL = '/swagger'
API_URL = '/static/swagger.yaml'
app.secret_key = b'a secret key'
SWAGGERUI_BLUEPRINT = get_swaggerui_blueprint(
    SWAGGER_URL,
    API_URL,
    config={
        'app_name': "Moose Detector"
    }
)
app.register_blueprint(SWAGGERUI_BLUEPRINT, url_prefix=SWAGGER_URL)
app.config['FLASK_ADMIN_SWATCH'] = 'readable'
UPLOAD_FOLDER = 'photos_buffer'
ALLOWED_EXTENSIONS_IMAGE = {'png', 'jpg', 'jpeg'}
ALLOWED_EXTENSIONS_VIDEO = {'mp4', 'webm', 'mov'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
babel = Babel(app)


def allowed_image_file(filename):
    return '.' in filename and \
        filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS_IMAGE


def allowed_image_video(filename):
    return '.' in filename and \
        filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS_VIDEO


class PredictionModelView(ModelView):
    page_size = 50
    can_create = False
    can_edit = False
    can_delete = False
    column_filters = ['camera_id', 'prediction_time', 'lynx_count', 'moose_count',
                      'wild_boar_count', 'brown_bear_count', 'other_count']

    def _list_thumbnail(self, context, model, name):
        return Markup(
            '<img src="data:image/jpg;base64, ' +
            codecs.encode(model.photo, encoding='base64').decode('utf-8') + '" width="500"/>'
        )

    column_formatters = {
        'photo': _list_thumbnail
    }


@app.route('/upload', methods=['GET', 'POST'])
def upload_file():
    print("in upload file")
    if request.method == 'POST':
        print("in upload post")
        if 'file' not in request.files:
            flash('Вы не загрузили файл', 'error')
        file = request.files['file']
        if file:
            if allowed_image_file(file.filename):
                process_photo(Image.open(file), request.form['camera_id'])
                preds = stop_processing(request.form['camera_id'],
                                       request.form['area_type'],
                                       request.form['prediction_date'])
                for pred in preds:
                    if insert_prediction(Prediction(**pred)):
                        flash('Файл загружен', 'success')
                    else:
                        flash('Кокая-то ошипка, попробуйте потом, попейте чаюб', 'error')
            elif allowed_image_video(file.filename):
                pred = process_video(file.stream, request.form['camera_id'], request.form['area_type'])
                if insert_prediction(Prediction(**pred)):
                    flash('Файл загружен', 'success')
                else:
                    flash('Кокая-то ошипка, попробуйте потом, попейте чаюб', 'error')
            else:
                flash('Недопустимый формат файла', 'error')
        else:
            flash('Файл пустой или не существует', 'error')
        return redirect('/upload')
    return redirect('admin/upload')


class UploadView(BaseView):
    @expose('/')
    def index(self):
        return self.render('upload_index.html')


class Prediction(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    prediction_time = db.Column(db.DateTime, nullable=False)
    camera_id = db.Column(db.Integer, nullable=False)
    photo = db.Column(db.LargeBinary(length=(2 ** 32) - 1), nullable=False)
    lynx_count = db.Column(db.Integer, nullable=False, default=0)
    brown_bear_count = db.Column(db.Integer, nullable=False, default=0)
    moose_count = db.Column(db.Integer, nullable=False, default=0)
    wild_boar_count = db.Column(db.Integer, nullable=False, default=0)
    other_count = db.Column(db.Integer, nullable=False, default=0)
    area_type = db.Column(db.String, nullable=False)

    def as_dict(self):
        return {
            'id': self.id,
            'prediction_time': self.prediction_time,
            'camera_id': self.camera_id,
            'photo': codecs.encode(self.photo, encoding='base64').decode('utf-8'),
            'lynx_count': self.lynx_count,
            'brown_bear_count': self.brown_bear_count,
            'moose_count': self.moose_count,
            'wild_boar_count': self.wild_boar_count,
            'other_count': self.other_count,
            'area_type': self.area_type
        }

    @classmethod
    def jsonify_all(cls):
        return jsonify(predictions=[pred.as_dict() for pred in cls.query.all()])


@app.route('/download/<filename>')
def download_file(filename):
    return send_file(filename, as_attachment=True)


class ReportView(BaseView):
    @expose('/')
    def index(self):
        camera_ids = get_cameras_ids()
        return self.render('report_index.html', camera_ids=camera_ids)


def get_cameras_ids():
    camera_ids = list(db.session.query(
        Prediction.camera_id
    ).distinct().all())
    camera_ids = [int(str(s)[1:-2]) for s in camera_ids]
    return camera_ids


@app.route('/report', methods=['POST', 'GET'])
def make_report():
    if request.method == 'POST':
        date_from = request.form['report_start_date']
        date_to = request.form['report_end_date']
        log(f"date_from:{date_from}, date_to:{date_to}")
        camera_ids = request.form.getlist('camera_ids')

        if not camera_ids:
            flash("Нет записей с камер в БД", "error")
            return redirect("admin/report")

        log(camera_ids)

        moose_forest = get_animal_in_area_count("moose", "forest", date_from, date_to, camera_ids)
        moose_swamp = get_animal_in_area_count("moose", "swamp", date_from, date_to, camera_ids)
        moose_field = get_animal_in_area_count("moose", "field", date_from, date_to, camera_ids)
        boar_forest = get_animal_in_area_count("wild_boar", "forest", date_from, date_to, camera_ids)
        boar_swamp = get_animal_in_area_count("wild_boar", "swamp", date_from, date_to, camera_ids)
        boar_field = get_animal_in_area_count("wild_boar", "field", date_from, date_to, camera_ids)
        total_moose = moose_forest + moose_swamp + moose_field
        total_boar = boar_forest + boar_field + boar_swamp

        wb = load_workbook('report.xlsm', keep_vba=True)
        ws = wb['Итоговый отчет по ЗВЕРЯМ']
        ws['X14'] = boar_forest
        ws['Y14'] = boar_field
        ws['Z14'] = boar_swamp
        ws['AA14'] = total_boar
        ws['X26'] = moose_forest
        ws['Y26'] = moose_field
        ws['Z26'] = moose_swamp
        ws['AA26'] = total_moose
        log(f"кабаны - лес:{boar_forest} поле:{boar_field} болото:{boar_swamp}, все:{total_boar}")
        log(f"лоси - лес:{moose_forest} поле:{moose_field} болото:{moose_swamp} все:{total_moose}")

        wb.save('report_new.xlsm')
        wb.close()

        return download_file('report_new.xlsm')
    return redirect('admin/report')


def get_animal_in_area_count(animal, area, start_date, end_date, cameras):
    query = db.session.query(func.sum(getattr(Prediction, f"{animal}_count")))\
        .filter(and_(Prediction.area_type == area,
                     Prediction.prediction_time.between(start_date, end_date),
                     Prediction.camera_id.in_(cameras)))
    result = query.scalar()
    return int(result) if result else 0


def to_date(date_string):
    return datetime.strptime(date_string, "%Y-%m-%d").date()


@app.route('/results')
# @auth.login_required
def get_response():
    date_to = request.args.get('dateTo', default=datetime.now().strftime("%Y-%m-%d"), type=to_date)
    date_from = request.args.get('dateFrom', default='2023-02-18', type=to_date)
    camera_id = request.args.get('cameraId', default=1, type=int)
    data = Prediction.query.filter_by(camera_id=camera_id).filter(
        Prediction.prediction_time.between(date_from, date_to)).all()
    return jsonify([pred.as_dict() for pred in data])


@app.route('/all')
# @auth.login_required
def get_all_data():
    return Prediction.jsonify_all()


@app.route('/photo/<camera_id>', methods=['GET', 'POST'])
# @auth.login_required
def insert_picture(camera_id):
    try:
        prediction_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with open('moose.jpg', 'rb') as file:
            photo = file.read()
        new_row = Prediction(camera_id=camera_id, prediction_time=prediction_time,
                             photo=photo, moose_count=1)
        db.session.add(new_row)
        db.session.commit()
        return make_response('', 204)
    except exc.SQLAlchemyError:
        return make_response('', 502)


def insert_prediction(prediction):
    try:
        db.session.add(prediction)
        db.session.commit()
        return True
    except exc.SQLAlchemyError as e:
        print('Sorry Error while inserting')
        print(e)
        log('INSERT PREDICITION\n' + str(e))
        return False


@auth.verify_password
def authenticate(username, password):
    if username and password:
        if username == os.environ['USER'] and password == os.environ['PASSWORD']:
            return True
        else:
            return False
    return False


if __name__ == "__main__":
    admin = Admin(app, name='moose_detector', template_mode='bootstrap3')
    admin.add_view(PredictionModelView(Prediction, db.session))
    admin.add_view(UploadView(name='Upload', endpoint='upload'))
    admin.add_view(ReportView(name='Report', endpoint='report'))
    app.run(host="0.0.0.0", port=5000)
